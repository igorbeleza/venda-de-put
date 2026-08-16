"""Extract fixtures from carteira_venda_put Excel for tests and seed data.

The .xlsx is gitignored (local reference). Copy it to the repo root to run this.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "carteira_venda_put (4).xlsx"
SHEET_ATIVOS = "Ativos Líquidos e Informações"

# Column letters → field names (1-indexed openpyxl columns)
ATIVO_FIELDS: list[tuple[int, str]] = [
    (1, "ticker"),  # A
    (2, "grupo"),  # B
    (3, "pl"),  # C
    (4, "pvp"),  # D
    (5, "dy"),  # E
    (6, "ev_ebitda"),  # F
    (7, "mrg_liq"),  # G
    (8, "liq_corr"),  # H
    (9, "roic"),  # I
    (10, "roe"),  # J
    (11, "div_pat"),  # K
    (12, "cresc"),  # L
    (13, "preco"),  # M
    (14, "mm200"),  # N
    (15, "ifr"),  # O
    (17, "boll_inf"),  # Q
    (24, "sinal"),  # X
    (25, "nroe"),  # Y
    (26, "nroic"),  # Z
    (27, "nmrgl"),  # AA
    (28, "ndiv"),  # AB
    (29, "nliqc"),  # AC
    (30, "npl"),  # AD
    (31, "npvp"),  # AE
    (32, "neveb"),  # AF
    (33, "ncrsc"),  # AG
    (38, "scoref"),  # AL
    (39, "pctf"),  # AM
    (41, "scoret"),  # AO
    (43, "scorec"),  # AQ
]


def _jsonable(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    return value


def extract_ativos(wb) -> list[dict]:
    ws = wb[SHEET_ATIVOS]
    ativos: list[dict] = []
    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row, 1).value
        if not ticker:
            continue
        item = {}
        for col, name in ATIVO_FIELDS:
            item[name] = _jsonable(ws.cell(row, col).value)
        ativos.append(item)
    return ativos


def extract_dados(wb) -> list[dict]:
    ws = wb["Dados"]
    rows: list[dict] = []
    for row in range(2, ws.max_row + 1):
        papel = ws.cell(row, 1).value
        if not papel:
            continue
        colunas = [_jsonable(ws.cell(row, c).value) for c in range(1, 23)]
        rows.append({"papel": papel, "colunas": colunas})
    return rows


def extract_feriados(wb) -> list[dict]:
    ws = wb["Feriados"]
    feriados: list[dict] = []
    for row in range(3, ws.max_row + 1):
        raw = ws.cell(row, 1).value
        if raw is None:
            continue
        if isinstance(raw, datetime):
            d = raw.date().isoformat()
        elif isinstance(raw, date):
            d = raw.isoformat()
        else:
            d = str(raw)[:10]
        descricao = ws.cell(row, 2).value or ""
        feriados.append({"date": d, "descricao": descricao})
    return feriados


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Excel not found: {XLSX}")

    wb = load_workbook(XLSX, data_only=True)
    ativos = extract_ativos(wb)
    dados = extract_dados(wb)
    feriados = extract_feriados(wb)
    universe = {a["ticker"]: a["grupo"] for a in ativos}

    fixtures_dir = ROOT / "tests" / "fixtures"
    fixtures_dir.mkdir(parents=True, exist_ok=True)
    data_dir = ROOT / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    (fixtures_dir / "excel_ativos.json").write_text(
        json.dumps(ativos, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (fixtures_dir / "excel_dados.json").write_text(
        json.dumps(dados, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (data_dir / "universe.json").write_text(
        json.dumps(universe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (data_dir / "feriados.json").write_text(
        json.dumps(feriados, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    brsr6 = next(a for a in ativos if a["ticker"] == "BRSR6")
    print(f"ativos: {len(ativos)}")
    print(f"dados rows: {len(dados)}")
    print(f"feriados: {len(feriados)}")
    print(f"universe: {len(universe)}")
    print(f"BRSR6 scoref={brsr6['scoref']!r} pctf={brsr6['pctf']!r}")


if __name__ == "__main__":
    main()
